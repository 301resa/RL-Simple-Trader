"""
evaluation/test_fold.py
========================
Standalone test-fold runner.

Loads every checkpoint saved during a training fold, runs deterministic
episodes on the test date range, prints a ranked summary table, and
generates a per-checkpoint Plotly HTML with:
  • Candlestick chart — entry ▲/▼ markers, coloured by win/loss
  • SL and TP horizontal lines per trade
  • Equity-curve subplot

Usage
-----
    python -m evaluation.test_fold \\
        --models-dir  logs/walk_forward/fold_00/models \\
        --config      config/ \\
        --data        data/ \\
        [--test-start 2024-01-01] \\
        [--test-end   2024-06-30] \\
        [--out-dir    logs/walk_forward/fold_00/test_results]

If --test-start / --test-end are omitted the script uses the DataSplitter
default test split (the held-out block after val).
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# ── Plotly (optional but expected) ───────────────────────────────────────────
try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    PLOTLY_OK = True
except ImportError:
    PLOTLY_OK = False


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _checkpoint_base(p: Path) -> str:
    """Return the base name of a checkpoint without the .zip extension."""
    name = p.name
    return name[:-4] if name.endswith(".zip") else name


def _find_checkpoints(models_dir: Path) -> List[Path]:
    """
    Return all checkpoint and hotsave files sorted by step number.

    Scans models_dir/ and models_dir/hotsaves/ (if present).
    Picks up:
      - checkpoint_*.zip   (eval-gated saves)
      - hotsave_*.zip      (training hot-saves)
      - legacy extension-less files (SB3 stripped the score decimal)
    """
    search_dirs = [models_dir]
    hotsaves_dir = models_dir / "hotsaves"
    if hotsaves_dir.is_dir():
        search_dirs.append(hotsaves_dir)

    found: List[Path] = []
    for d in search_dirs:
        # .zip files — both checkpoint_ and hotsave_ prefixes
        zips = [
            f for f in d.glob("*.zip")
            if f.stem.startswith(("checkpoint_", "hotsave_"))
        ]
        if zips:
            found.extend(zips)
        else:
            # Legacy: SB3 stripped the score decimal as a file extension
            found.extend(
                f for f in d.iterdir()
                if not f.name.endswith(".pkl")
                and not f.is_dir()
                and (f.stem.startswith("checkpoint_") or f.stem.startswith("hotsave_"))
            )

    if not found:
        raise FileNotFoundError(f"No checkpoint or hotsave files found in {models_dir}")

    return sorted(set(found), key=_step_from_name)


def _vecnorm_path(ckpt: Path) -> Optional[Path]:
    """Derive paired VecNormalize .pkl path from checkpoint path."""
    base = _checkpoint_base(ckpt)   # strip .zip if present
    candidate = ckpt.parent / f"{base}_vecnormalize.pkl"
    if candidate.exists():
        return candidate
    # Fallback: best_model_vecnormalize.pkl or vecnormalize.pkl in same dir
    for name in ("best_model_vecnormalize.pkl", "vecnormalize.pkl"):
        fb = ckpt.parent / name
        if fb.exists():
            return fb
    return None


def _composite_from_name(p: Path) -> float:
    """Parse composite score from checkpoint filename, or 0 if FINAL_STEP."""
    base = _checkpoint_base(p)
    if "_c" in base:
        try:
            return float(base.split("_c")[-1])
        except ValueError:
            pass
    return 0.0


def _step_from_name(p: Path) -> int:
    """Parse training step from checkpoint or hotsave filename."""
    base = _checkpoint_base(p)
    # checkpoint_sNN_stepNNNNNN_... format
    for token in base.split("_"):
        if token.startswith("step"):
            try:
                return int(token[4:])
            except ValueError:
                pass
    # FINAL_STEP format
    if "FINAL" in base:
        part = base.split("FINAL_STEP")[-1]
        try:
            return int(part)
        except ValueError:
            pass
    # hotsave_NNNNNNNNNN format — last numeric token
    tokens = base.split("_")
    for token in reversed(tokens):
        if token.isdigit():
            return int(token)
    return 0


# ─────────────────────────────────────────────────────────────────────────────
# Rollout
# ─────────────────────────────────────────────────────────────────────────────

def _run_checkpoint(
    ckpt_path: Path,
    vn_path: Optional[Path],
    env_factory,
    n_episodes: int = 20,
    n_workers: int = 8,
) -> Tuple[List[dict], List[dict]]:
    """
    Load checkpoint + VecNormalize, run n_episodes deterministically.

    Uses SubprocVecEnv with n_workers parallel envs so episodes execute
    in parallel rather than one-by-one, matching training throughput.

    Returns
    -------
    trades   : list of trade dicts (raw, from _episode_summary)
    episodes : list of episode-level summary dicts
    """
    import multiprocessing
    from stable_baselines3.common.vec_env import SubprocVecEnv, VecNormalize
    from sb3_contrib import RecurrentPPO

    n_workers = min(n_workers, n_episodes, multiprocessing.cpu_count())

    vec_env = SubprocVecEnv([env_factory] * n_workers, start_method="spawn")
    vec_env = VecNormalize(vec_env, norm_obs=True, norm_reward=False,
                           clip_obs=10.0, training=False)

    if vn_path is not None:
        vec_env = VecNormalize.load(str(vn_path), vec_env)
        vec_env.training = False
        vec_env.norm_reward = False

    model = RecurrentPPO.load(str(ckpt_path), env=vec_env)

    all_trades: List[dict] = []
    all_episodes: List[dict] = []
    completed = 0

    obs = vec_env.reset()
    lstm_states = None
    episode_starts = np.ones((n_workers,), dtype=bool)

    while completed < n_episodes:
        action, lstm_states = model.predict(
            obs,
            state=lstm_states,
            episode_start=episode_starts,
            deterministic=True,
        )
        obs, _reward, dones, infos = vec_env.step(action)
        episode_starts = dones.copy()

        for done, info in zip(dones, infos):
            if done and info:
                all_episodes.append(info)
                all_trades.extend(info.get("trades_list", []))
                completed += 1
                if completed >= n_episodes:
                    break

    vec_env.close()
    return all_trades, all_episodes


# ─────────────────────────────────────────────────────────────────────────────
# Metrics
# ─────────────────────────────────────────────────────────────────────────────

def _compute_metrics(trades: List[dict], episodes: List[dict]) -> dict:
    """Compute summary metrics from raw trade list."""
    if not trades:
        return {
            "n_trades": 0, "win_rate": 0.0,
            "total_pnl_r": 0.0, "total_pnl_dollars": 0.0,
            "avg_win_r": 0.0, "avg_loss_r": 0.0,
            "avg_win_dollars": 0.0, "avg_loss_dollars": 0.0,
            "profit_factor": 0.0, "profit_factor_usd": 0.0, "sharpe": 0.0,
            "max_dd_r": 0.0, "max_dd_dollars": 0.0,
            "avg_duration_min": 0.0,
        }

    pnl_r   = [t["pnl_r"] for t in trades]
    pnl_usd = [t.get("pnl_dollars", 0.0) for t in trades]
    wins_r   = [p for p in pnl_r   if p > 0]
    losses_r = [p for p in pnl_r   if p <= 0]
    wins_u   = [p for p in pnl_usd if p > 0]
    losses_u = [p for p in pnl_usd if p <= 0]
    n = len(pnl_r)

    win_rate   = len(wins_r) / n
    total_pnl  = sum(pnl_r)
    total_usd  = sum(pnl_usd)
    avg_win    = float(np.mean(wins_r))   if wins_r   else 0.0
    avg_loss   = float(abs(np.mean(losses_r))) if losses_r else 0.0
    avg_win_u  = float(np.mean(wins_u))   if wins_u   else 0.0
    avg_loss_u = float(abs(np.mean(losses_u))) if losses_u else 0.0
    # R-based profit factor (kept for reference)
    gp = sum(wins_r)   if wins_r   else 0.0
    gl = abs(sum(losses_r)) if losses_r else 1e-9
    pf = gp / gl

    # Dollar-based profit factor — accounts for position sizing variation.
    # This is the honest metric: a high-confluence loss (large position) can
    # make dollar PnL negative even when R-PF looks positive.
    gp_u = sum(wins_u)         if wins_u   else 0.0
    gl_u = abs(sum(losses_u))  if losses_u else 1e-9
    pf_dollars = gp_u / gl_u

    # Sharpe: episode-level daily PnL std (annualised)
    ep_pnls = [ep.get("total_pnl_r", 0.0) for ep in episodes if ep]
    sharpe = 0.0
    if len(ep_pnls) >= 2:
        mu = float(np.mean(ep_pnls))
        sd = float(np.std(ep_pnls))
        sharpe = (mu / sd) * np.sqrt(252) if sd > 1e-9 else 0.0

    # Max drawdown (R and $)
    eq_r   = np.cumsum(pnl_r)
    eq_usd = np.cumsum(pnl_usd)
    peak_r   = np.maximum.accumulate(eq_r)
    peak_usd = np.maximum.accumulate(eq_usd)
    max_dd   = float((peak_r   - eq_r).max())   if n > 0 else 0.0
    max_dd_u = float((peak_usd - eq_usd).max()) if n > 0 else 0.0

    avg_dur = float(np.mean([t.get("duration_min", 0) for t in trades]))

    return {
        "n_trades":           n,
        "win_rate":           round(win_rate,    4),
        "total_pnl_r":        round(total_pnl,   4),
        "total_pnl_dollars":  round(total_usd,   2),
        "avg_win_r":          round(avg_win,      4),
        "avg_loss_r":         round(avg_loss,     4),
        "avg_win_dollars":    round(avg_win_u,    2),
        "avg_loss_dollars":   round(avg_loss_u,   2),
        "profit_factor":      round(pf,           4),   # R-based (reference)
        "profit_factor_usd":  round(pf_dollars,   4),   # dollar-based (primary)
        "sharpe":             round(sharpe,        4),
        "max_dd_r":           round(max_dd,        4),
        "max_dd_dollars":     round(max_dd_u,      2),
        "avg_duration_min":   round(avg_dur,       1),
    }


def _composite(m: dict,
               w_sharpe: float = 0.30,
               w_pnl: float = 0.30,
               w_wl: float = 0.20,
               w_dd: float = 0.20) -> float:
    """
    Composite ranking score — dollar-based primary metrics.

    Design:
      - Uses dollar PF and dollar PnL as primary profit signals so that
        position-size variation (confluence grading) is fully accounted for.
      - Hard gates ensure losing models (negative dollar PnL, dollar PF < 1,
        or fewer than 20 trades) always rank below every profitable model.
      - Profitable models score in [0.1, 1.0]; losing models in [0.0, 0.099].

    Weights:  Sharpe 30% | PnL$ 30% | W/L ratio 20% | MaxDD 20%
    """
    # ── Hard gate 1: insufficient trades — no statistical validity ────────────
    if m["n_trades"] < 20:
        return round(m["n_trades"] / 200.0 * 0.05, 4)  # tiny score, sub-ranks by volume

    # ── Hard gate 2: losing in dollars OR dollar PF < 1 ──────────────────────
    if m["total_pnl_dollars"] <= 0.0 or m["profit_factor_usd"] < 1.0:
        pf_sub = min(m["profit_factor_usd"], 1.0) * 0.05
        return round(max(0.0, pf_sub), 4)

    # ── Profitable models: full composite ─────────────────────────────────────
    sharpe  = max(m["sharpe"], 0.0)
    pnl_usd = m["total_pnl_dollars"]
    # W/L ratio using dollar averages (accounts for sizing)
    wl = m["win_rate"] * (m["avg_win_dollars"] / max(m["avg_loss_dollars"], 1e-6))
    dd = m["max_dd_dollars"]

    # Reference targets
    ref_sharpe  = 1.5      # annualised Sharpe
    ref_pnl_usd = 5000.0   # dollar PnL target over test period
    ref_wl      = 3.0      # win_rate × avg_win$/avg_loss$ target
    ref_dd_usd  = 3000.0   # max drawdown $ target

    sn = min(sharpe   / ref_sharpe,  1.0)
    pn = min(pnl_usd  / ref_pnl_usd, 1.0)
    wn = min(wl       / ref_wl,      1.0)
    dn = min(dd       / ref_dd_usd,  1.0)

    total_w = w_sharpe + w_pnl + w_wl + w_dd
    raw = (w_sharpe * sn + w_pnl * pn + w_wl * wn + w_dd * (1.0 - dn)) / total_w
    return round(max(raw, 0.1), 4)


# ─────────────────────────────────────────────────────────────────────────────
# Console tables
# ─────────────────────────────────────────────────────────────────────────────

_COL_W = {
    "ckpt":      28,
    "trades":     6,
    "wr":         7,
    "pnl":       10,   # dollars: e.g. $-12,345
    "pf_usd":     6,   # dollar-based PF (primary)
    "sharpe":     8,
    "dd":         9,   # dollars: e.g. $12,345
    "dur":        7,
    "comp":       7,
}

def _header_row() -> str:
    return (
        f"{'Checkpoint':<{_COL_W['ckpt']}}"
        f"{'Trades':>{_COL_W['trades']}}"
        f"{'WR%':>{_COL_W['wr']}}"
        f"{'PnL($)':>{_COL_W['pnl']}}"
        f"{'$PF':>{_COL_W['pf_usd']}}"
        f"{'Sharpe':>{_COL_W['sharpe']}}"
        f"{'MaxDD($)':>{_COL_W['dd']}}"
        f"{'AvgMin':>{_COL_W['dur']}}"
        f"{'Score':>{_COL_W['comp']}}"
    )


def _data_row(name: str, m: dict, score: float) -> str:
    short = name[:_COL_W["ckpt"] - 1] if len(name) > _COL_W["ckpt"] - 1 else name
    pnl_str = f"${m['total_pnl_dollars']:+,.0f}"
    dd_str  = f"${m['max_dd_dollars']:,.0f}"
    return (
        f"{short:<{_COL_W['ckpt']}}"
        f"{m['n_trades']:>{_COL_W['trades']}}"
        f"{m['win_rate']*100:>{_COL_W['wr']}.1f}"
        f"{pnl_str:>{_COL_W['pnl']}}"
        f"{m['profit_factor_usd']:>{_COL_W['pf_usd']}.2f}"
        f"{m['sharpe']:>{_COL_W['sharpe']}.3f}"
        f"{dd_str:>{_COL_W['dd']}}"
        f"{m['avg_duration_min']:>{_COL_W['dur']}.1f}"
        f"{score:>{_COL_W['comp']}.3f}"
    )


def _separator() -> str:
    return "-" * sum(_COL_W.values())


# ─────────────────────────────────────────────────────────────────────────────
# Per-model trading journal
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_trade_times(trades: List[dict], data_loader) -> List[dict]:
    """
    Add 'entry_time' and 'exit_time' ISO strings to each trade dict by
    mapping bar indices into the actual session timestamps from the data loader.

    entry_bar_idx / exit_bar_idx are indices into the session bars for that day.
    """
    enriched = []
    day_bar_cache: dict = {}

    for t in trades:
        date = t["date"]
        if date not in day_bar_cache:
            try:
                day_bar_cache[date] = data_loader.get_day_bars(date)
            except Exception:
                day_bar_cache[date] = pd.DataFrame()

        day_bars = day_bar_cache[date]
        n = len(day_bars)
        entry_idx = min(int(t.get("entry_bar_idx", 0)), max(n - 1, 0))
        exit_idx  = min(int(t.get("exit_bar_idx",  0)), max(n - 1, 0))

        tc = t.copy()
        if n > 0:
            tc["entry_time"] = str(day_bars.index[entry_idx])
            tc["exit_time"]  = str(day_bars.index[exit_idx])
        else:
            tc["entry_time"] = date + " 00:00"
            tc["exit_time"]  = date + " 00:00"
        enriched.append(tc)

    return enriched


def _build_journal(
    ckpt_name: str,
    trades: List[dict],
    metrics: dict,
    score: float,
    data_loader,
    test_days: List[str],
    out_path: Path,
) -> None:
    """
    Build a self-contained HTML trading journal for one model checkpoint.

    Layout
    ------
    1. Header — model name + test period
    2. Metrics card — all trading metrics in styled tiles
    3. Candlestick chart — full test period bars with entry/exit/SL/TP overlays
    4. Equity curve + drawdown chart
    5. Trade-by-trade table — sortable, win/loss colour coded
    """
    if not PLOTLY_OK:
        return

    trades = _resolve_trade_times(trades, data_loader)

    # ── Load all bars for the test period ─────────────────────────────────────
    bars_list = []
    for d in test_days:
        try:
            db = data_loader.get_day_bars(d)
            if db is not None and not db.empty:
                bars_list.append(db)
        except Exception:
            pass

    has_bars = bool(bars_list)
    if has_bars:
        all_bars = pd.concat(bars_list).sort_index().reset_index()
        time_col = all_bars.columns[0]

    # ── Build candlestick + PnL chart ─────────────────────────────────────────
    # Rows 2 & 3 use trade-index x-axis (not shared) so bars always fill width
    fig = make_subplots(
        rows=3, cols=1,
        shared_xaxes=False,
        row_heights=[0.55, 0.22, 0.23],
        vertical_spacing=0.05,
        subplot_titles=["Price & Trades", "Per-Trade PnL (R)", "Cumulative PnL (R)"],
    )

    if has_bars:
        fig.add_trace(
            go.Candlestick(
                x=all_bars[time_col],
                open=all_bars["open"],
                high=all_bars["high"],
                low=all_bars["low"],
                close=all_bars["close"],
                name="Price",
                increasing_line_color="#26a69a",
                decreasing_line_color="#ef5350",
                showlegend=False,
            ),
            row=1, col=1,
        )

    # ── Trade overlays ────────────────────────────────────────────────────────
    pnl_times: list = []
    pnl_vals:  List[float] = []

    for t in trades:
        is_long    = t["direction"].upper() == "LONG"
        is_win     = t["is_win"]
        entry_px   = t["entry_price"]
        stop_px    = t["stop_price"]
        tgt_px     = t["initial_target"]
        exit_px    = t["exit_price"]
        pnl_r      = t["pnl_r"]
        entry_time = t["entry_time"]
        exit_time  = t["exit_time"]
        reason     = t.get("exit_reason", "")
        dur_min    = t.get("duration_min", 0)
        contracts  = t.get("n_contracts", 1)
        pnl_pts    = t.get("pnl_points", 0.0)

        entry_sym = "triangle-up"   if is_long else "triangle-down"
        entry_col = "#2196F3"
        exit_col  = "#26a69a"       if is_win  else "#ef5350"

        hover_entry = (
            f"<b>{'LONG ▲' if is_long else 'SHORT ▼'} Entry</b><br>"
            f"Time : {entry_time}<br>"
            f"Price: {entry_px}<br>"
            f"SL   : {stop_px}<br>"
            f"TP   : {tgt_px}<br>"
            f"Lots : {contracts}<br>"
            f"PnL  : {pnl_r:+.2f}R  ({pnl_pts:+.2f} pts)<extra></extra>"
        )
        hover_exit = (
            f"<b>Exit — {reason}</b><br>"
            f"Time : {exit_time}<br>"
            f"Price: {exit_px}<br>"
            f"PnL  : {pnl_r:+.2f}R<br>"
            f"Dur  : {dur_min} min<extra></extra>"
        )

        # Entry marker
        fig.add_trace(go.Scatter(
            x=[entry_time], y=[entry_px], mode="markers",
            marker=dict(symbol=entry_sym, color=entry_col, size=13,
                        line=dict(width=1, color="white")),
            showlegend=False, hovertemplate=hover_entry,
        ), row=1, col=1)

        # Exit marker
        fig.add_trace(go.Scatter(
            x=[exit_time], y=[exit_px], mode="markers",
            marker=dict(symbol="x", color=exit_col, size=11,
                        line=dict(width=2, color=exit_col)),
            showlegend=False, hovertemplate=hover_exit,
        ), row=1, col=1)

        # Entry→exit connector (thin white dashed line)
        fig.add_trace(go.Scatter(
            x=[entry_time, exit_time], y=[entry_px, exit_px],
            mode="lines",
            line=dict(color="rgba(255,255,255,0.25)", width=1, dash="dot"),
            showlegend=False, hoverinfo="skip",
        ), row=1, col=1)

        # SL line (solid red, entry→exit) — use Scatter so it renders on the correct subplot
        fig.add_trace(go.Scatter(
            x=[entry_time, exit_time], y=[stop_px, stop_px],
            mode="lines",
            line=dict(color="rgba(239,83,80,0.80)", width=1.5, dash="dash"),
            showlegend=False,
            hovertemplate=f"<b>Stop Loss</b>: {stop_px}<extra></extra>",
        ), row=1, col=1)

        # TP line (solid green, entry→exit) — same reason
        fig.add_trace(go.Scatter(
            x=[entry_time, exit_time], y=[tgt_px, tgt_px],
            mode="lines",
            line=dict(color="rgba(38,166,154,0.80)", width=1.5, dash="dash"),
            showlegend=False,
            hovertemplate=f"<b>Take Profit</b>: {tgt_px}<extra></extra>",
        ), row=1, col=1)

        pnl_times.append(exit_time)
        pnl_vals.append(pnl_r)

    # Trade index x-axis for bottom chart so bars always fill the width
    trade_idx  = list(range(1, len(pnl_vals) + 1))
    cum_pnl    = list(np.cumsum(pnl_vals)) if pnl_vals else []
    bar_colors = ["#26a69a" if v >= 0 else "#ef5350" for v in pnl_vals]
    cum_color  = "#26a69a" if (cum_pnl[-1] if cum_pnl else 0) >= 0 else "#ef5350"

    # Per-trade PnL bars — row 2
    fig.add_trace(go.Bar(
        x=trade_idx,
        y=pnl_vals,
        marker_color=bar_colors,
        marker_line_width=0,
        name="Trade PnL",
        showlegend=False,
        hovertemplate="Trade #%{x}<br>PnL: %{y:+.2f}R<extra></extra>",
    ), row=2, col=1)
    fig.add_hline(y=0, line=dict(color="rgba(255,255,255,0.25)", width=1, dash="dash"), row=2, col=1)

    # Cumulative PnL line — row 3
    fig.add_trace(go.Scatter(
        x=trade_idx,
        y=cum_pnl,
        mode="lines",
        line=dict(color=cum_color, width=2),
        fill="tozeroy",
        fillcolor="rgba(38,166,154,0.12)" if cum_color == "#26a69a" else "rgba(239,83,80,0.12)",
        name="Cumulative PnL",
        showlegend=False,
        hovertemplate="Trade #%{x}<br>Cumulative: %{y:+.2f}R<extra></extra>",
    ), row=3, col=1)
    fig.add_hline(y=0, line=dict(color="rgba(255,255,255,0.25)", width=1, dash="dash"), row=3, col=1)

    fig.update_layout(
        template="plotly_dark",
        height=1050,
        dragmode="pan",
        margin=dict(l=60, r=20, t=40, b=40),
    )
    fig.update_yaxes(title_text="Price",              row=1, col=1)
    fig.update_yaxes(title_text="PnL (R)",            row=2, col=1)
    fig.update_yaxes(title_text="Cumulative PnL (R)", row=3, col=1)
    fig.update_xaxes(rangeslider=dict(visible=False), row=1, col=1)
    fig.update_xaxes(rangeslider=dict(visible=False), row=2, col=1)
    fig.update_xaxes(title_text="Trade #",
                     rangeslider=dict(visible=True, thickness=0.04, bgcolor="#1e222d"),
                     row=3, col=1)

    chart_html = fig.to_html(
        full_html=False,
        include_plotlyjs="cdn",
        config={"scrollZoom": True, "displayModeBar": True,
                "modeBarButtonsToAdd": ["pan2d", "zoom2d"]},
    )

    # ── Metrics card ──────────────────────────────────────────────────────────
    m = metrics
    wr_pct    = m["win_rate"] * 100
    wl_ratio  = m["avg_win_r"] / max(m["avg_loss_r"], 1e-6)

    def tile(label: str, value: str, cls: str = "") -> str:
        return (
            f'<div class="tile">'
            f'<div class="tlabel">{label}</div>'
            f'<div class="tvalue {cls}">{value}</div>'
            f'</div>'
        )

    pos = "pos" if m["total_pnl_r"] >= 0 else "neg"
    tiles = "".join([
        tile("Trades",       str(m["n_trades"])),
        tile("Win Rate",     f"{wr_pct:.1f}%",                          "pos" if wr_pct >= 50 else "neg"),
        tile("W/L Ratio",    f"{wl_ratio:.2f}",                         "pos" if wl_ratio >= 1 else "neg"),
        tile("Total PnL",    f"${m['total_pnl_dollars']:+,.0f}",        pos),
        tile("PnL (R)",      f"{m['total_pnl_r']:+.2f}R",               pos),
        tile("PF ($)",       f"{m['profit_factor_usd']:.2f}",            "pos" if m['profit_factor_usd'] >= 1 else "neg"),
        tile("Sharpe",       f"{m['sharpe']:.3f}",                      "pos" if m['sharpe'] >= 0 else "neg"),
        tile("Max DD",       f"${m['max_dd_dollars']:,.0f}",            "neg" if m['max_dd_dollars'] > 500 else "pos"),
        tile("Avg Win",      f"${m['avg_win_dollars']:,.0f}",           "pos"),
        tile("Avg Loss",     f"-${m['avg_loss_dollars']:,.0f}",         "neg"),
        tile("Avg Dur",      f"{m['avg_duration_min']:.0f} min"),
        tile("Score",        f"{score:.3f}",                             "pos" if score >= 0.2 else "neg"),
    ])

    # ── Trade table ───────────────────────────────────────────────────────────
    rows_html = ""
    for i, t in enumerate(sorted(trades, key=lambda x: x["entry_time"]), 1):
        cls = "win-row" if t["is_win"] else "loss-row"
        pnl_cls = "pos" if t["pnl_r"] >= 0 else "neg"
        rows_html += (
            f'<tr class="{cls}">'
            f'<td>{i}</td>'
            f'<td>{t["date"]}</td>'
            f'<td>{"▲ LONG" if t["direction"].upper()=="LONG" else "▼ SHORT"}</td>'
            f'<td>{t["entry_time"][11:16] if len(t["entry_time"]) > 10 else t["entry_time"]}</td>'
            f'<td>{t["exit_time"][11:16]  if len(t["exit_time"])  > 10 else t["exit_time"]}</td>'
            f'<td>{t["duration_min"]} min</td>'
            f'<td>{t["entry_price"]}</td>'
            f'<td>{t["exit_price"]}</td>'
            f'<td style="color:#ef5350">{t["stop_price"]}</td>'
            f'<td style="color:#26a69a">{t["initial_target"]}</td>'
            f'<td>{t.get("n_contracts", 1)}</td>'
            f'<td class="{pnl_cls}">{t["pnl_r"]:+.2f}</td>'
            f'<td class="{pnl_cls}">{t.get("pnl_points", 0.0):+.2f}</td>'
            f'<td class="{pnl_cls}">${t.get("pnl_dollars", 0.0):+.0f}</td>'
            f'<td>{t.get("exit_reason", "")}</td>'
            f'<td>{t.get("mae_r", 0.0):.2f}</td>'
            f'</tr>\n'
        )

    table_html = f"""
<table id="trade-table">
  <thead>
    <tr>
      <th>#</th><th>Date</th><th>Dir</th><th>Entry</th><th>Exit</th>
      <th>Dur</th><th>Entry Px</th><th>Exit Px</th><th>SL</th><th>TP</th>
      <th>Lots</th><th>PnL (R)</th><th>PnL (pts)</th><th>PnL ($)</th>
      <th>Reason</th><th>MAE (R)</th>
    </tr>
  </thead>
  <tbody>
    {rows_html}
  </tbody>
</table>"""

    # ── Assemble full HTML ────────────────────────────────────────────────────
    period_str = f"{test_days[0]} → {test_days[-1]}  ({len(test_days)} days)"
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{ckpt_name}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #131722; color: #d1d4dc; font-family: 'Segoe UI', monospace; font-size: 13px; }}
  h1 {{ padding: 16px 20px 4px; font-size: 16px; font-weight: 600; color: #e0e0e0; }}
  .subtitle {{ padding: 0 20px 12px; color: #888; font-size: 12px; }}
  .tiles {{ display: flex; flex-wrap: wrap; gap: 10px; padding: 12px 20px; }}
  .tile {{ background: #1e222d; border: 1px solid #2a2e39; border-radius: 6px;
           padding: 10px 18px; min-width: 100px; text-align: center; }}
  .tlabel {{ font-size: 10px; color: #666; text-transform: uppercase; letter-spacing: .5px; margin-bottom: 4px; }}
  .tvalue {{ font-size: 20px; font-weight: 700; }}
  .pos {{ color: #26a69a; }}
  .neg {{ color: #ef5350; }}
  .chart-wrap {{ padding: 0 10px; }}
  h2 {{ padding: 16px 20px 8px; font-size: 13px; color: #aaa; text-transform: uppercase; letter-spacing: 1px; border-top: 1px solid #2a2e39; margin-top: 12px; }}
  table {{ width: calc(100% - 40px); margin: 0 20px 30px; border-collapse: collapse; font-size: 12px; }}
  thead tr {{ background: #1e222d; }}
  th {{ padding: 8px 10px; text-align: right; color: #888; font-weight: 600;
        border-bottom: 2px solid #2a2e39; white-space: nowrap; }}
  th:nth-child(-n+3) {{ text-align: left; }}
  td {{ padding: 6px 10px; border-bottom: 1px solid #1e222d; text-align: right; }}
  td:nth-child(-n+3) {{ text-align: left; }}
  tr.win-row  td {{ background: rgba(38,166,154,0.04); }}
  tr.loss-row td {{ background: rgba(239,83,80,0.04); }}
  tr:hover td {{ background: rgba(255,255,255,0.04) !important; }}
</style>
</head>
<body>
<h1>&#x1F4C8; {ckpt_name}</h1>
<div class="subtitle">Test period: {period_str}</div>
<div class="tiles">{tiles}</div>
<div class="chart-wrap">{chart_html}</div>
<h2>Trade Journal — {len(trades)} trades</h2>
{table_html}
</body>
</html>"""

    out_path.write_text(html, encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# Excel journal per model
# ─────────────────────────────────────────────────────────────────────────────

def _build_excel_journal(
    ckpt_name: str,
    trades: List[dict],
    metrics: dict,
    score: float,
    test_days: List[str],
    out_path: Path,
) -> None:
    """
    Write a self-contained Excel workbook for one model checkpoint.

    Sheets
    ------
    Trades   — one row per trade, all fields, green/red fill by win/loss
    Daily    — per-date aggregates (trades, wins, PnL R/$)
    Metrics  — full metric summary + score
    """
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    GREEN_FILL  = PatternFill("solid", fgColor="C8E6C9")
    RED_FILL    = PatternFill("solid", fgColor="FFCDD2")
    HEADER_FONT = Font(bold=True)
    CENTER      = Alignment(horizontal="center")

    _TRADE_COLS = [
        "date", "direction", "entry_time", "exit_time",
        "entry_price", "stop_price", "initial_target", "exit_price",
        "n_contracts", "duration_min", "exit_reason",
        "pnl_r", "pnl_points", "pnl_dollars", "is_win", "mae_r",
    ]

    def _auto_width(ws):
        for i, col in enumerate(ws.columns, 1):
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(i)].width = min(w + 2, 24)

    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:

        # ── Sheet 1: Trades ───────────────────────────────────────
        if trades:
            df = pd.DataFrame(trades)
            cols = [c for c in _TRADE_COLS if c in df.columns]
            # add entry/exit time columns if enriched by caller
            for extra in ("entry_time", "exit_time"):
                if extra in df.columns and extra not in cols:
                    cols.insert(2, extra)
            df_out = df[cols].copy()
        else:
            df_out = pd.DataFrame(columns=_TRADE_COLS)

        df_out.to_excel(writer, sheet_name="Trades", index=False)
        ws_t = writer.sheets["Trades"]
        for cell in ws_t[1]:
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        # find is_win column index for row fills
        is_win_col = next(
            (i for i, cell in enumerate(ws_t[1], 1) if cell.value == "is_win"),
            None,
        )
        for row in ws_t.iter_rows(min_row=2):
            fill = GREEN_FILL if (is_win_col and row[is_win_col - 1].value) else RED_FILL
            for cell in row:
                cell.fill = fill
        _auto_width(ws_t)

        # ── Sheet 2: Daily ────────────────────────────────────────
        if not df_out.empty and "date" in df_out.columns:
            daily = (
                df_out.groupby("date").agg(
                    trades=("pnl_r", "count"),
                    wins=("is_win", "sum"),
                    total_pnl_r=("pnl_r", "sum"),
                    total_pnl_dollars=("pnl_dollars", "sum") if "pnl_dollars" in df_out.columns else ("pnl_r", "sum"),
                    avg_duration_min=("duration_min", "mean") if "duration_min" in df_out.columns else ("pnl_r", "count"),
                ).reset_index()
            )
            daily["win_rate"] = (daily["wins"] / daily["trades"].clip(lower=1)).round(3)
            daily.to_excel(writer, sheet_name="Daily", index=False)
            ws_d = writer.sheets["Daily"]
            for cell in ws_d[1]:
                cell.font = HEADER_FONT
            pnl_col = next(
                (i for i, cell in enumerate(ws_d[1], 1) if cell.value == "total_pnl_r"),
                None,
            )
            for row in ws_d.iter_rows(min_row=2):
                v = row[pnl_col - 1].value if pnl_col else 0
                fill = GREEN_FILL if (v or 0) >= 0 else RED_FILL
                for cell in row:
                    cell.fill = fill
            _auto_width(ws_d)

        # ── Sheet 3: Metrics ──────────────────────────────────────
        period = f"{test_days[0]} → {test_days[-1]}" if test_days else ""
        metric_rows = [
            ("Model",              ckpt_name),
            ("Test Period",        period),
            ("Test Days",          len(test_days)),
            ("",                   ""),
            ("Trades",             metrics.get("n_trades", 0)),
            ("Win Rate",           f"{metrics.get('win_rate', 0)*100:.1f}%"),
            ("Total PnL (R)",      f"{metrics.get('total_pnl_r', 0):+.4f}"),
            ("Total PnL ($)",      f"${metrics.get('total_pnl_dollars', 0):+,.2f}"),
            ("Profit Factor (R)",  f"{metrics.get('profit_factor', 0):.4f}"),
            ("Profit Factor ($)",  f"{metrics.get('profit_factor_usd', 0):.4f}"),
            ("Sharpe Ratio",       f"{metrics.get('sharpe', 0):.4f}"),
            ("Max Drawdown (R)",   f"{metrics.get('max_dd_r', 0):.4f}"),
            ("Max Drawdown ($)",   f"${metrics.get('max_dd_dollars', 0):,.2f}"),
            ("Avg Win (R)",        f"{metrics.get('avg_win_r', 0):+.4f}"),
            ("Avg Loss (R)",       f"{metrics.get('avg_loss_r', 0):.4f}"),
            ("Avg Win ($)",        f"${metrics.get('avg_win_dollars', 0):,.2f}"),
            ("Avg Loss ($)",       f"${metrics.get('avg_loss_dollars', 0):,.2f}"),
            ("Avg Duration",       f"{metrics.get('avg_duration_min', 0):.1f} min"),
            ("",                   ""),
            ("Composite Score",    f"{score:.4f}"),
        ]
        mdf = pd.DataFrame(metric_rows, columns=["Metric", "Value"])
        mdf.to_excel(writer, sheet_name="Metrics", index=False)
        ws_m = writer.sheets["Metrics"]
        for cell in ws_m[1]:
            cell.font = HEADER_FONT
        for row in ws_m.iter_rows(min_row=2):
            row[0].font = Font(bold=True)
        ws_m.column_dimensions["A"].width = 24
        ws_m.column_dimensions["B"].width = 22



def _build_leaderboard_excel(
    results: List[Dict],
    test_days: List[str],
    out_path: Path,
) -> None:
    """
    Write a single leaderboard Excel with all models on one sheet,
    ranked best-first by composite score.  Top row highlighted green.
    """
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    GOLD_FILL   = PatternFill("solid", fgColor="FFF9C4")
    GREEN_FILL  = PatternFill("solid", fgColor="C8E6C9")
    RED_FILL    = PatternFill("solid", fgColor="FFCDD2")
    HEADER_FONT = Font(bold=True)

    sorted_r = sorted(results, key=lambda r: r["score"], reverse=True)

    rows = []
    for rank, r in enumerate(sorted_r, 1):
        m = r["metrics"]
        rows.append({
            "Rank":          rank,
            "Model":         r["name"],
            "Score":         r["score"],
            "Trades":        m["n_trades"],
            "WR%":           round(m["win_rate"] * 100, 1),
            "PnL_R":         m["total_pnl_r"],
            "PnL_$":         m["total_pnl_dollars"],
            "PF_R":          m["profit_factor"],
            "PF_$":          m["profit_factor_usd"],
            "Sharpe":        m["sharpe"],
            "MaxDD_R":       m["max_dd_r"],
            "MaxDD_$":       m["max_dd_dollars"],
            "AvgWin_R":      m["avg_win_r"],
            "AvgLoss_R":     m["avg_loss_r"],
            "AvgWin_$":      m["avg_win_dollars"],
            "AvgLoss_$":     m["avg_loss_dollars"],
            "AvgDur_min":    m["avg_duration_min"],
        })

    df = pd.DataFrame(rows)
    period = f"{test_days[0]} → {test_days[-1]}" if test_days else ""

    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        # ── Leaderboard sheet ─────────────────────────────────────
        df.to_excel(writer, sheet_name="Leaderboard", index=False, startrow=1)
        ws = writer.sheets["Leaderboard"]

        # Title row
        ws["A1"] = f"Test Fold Leaderboard  —  {period}  ({len(test_days)} days)"
        ws["A1"].font = Font(bold=True, size=12)

        # Header formatting
        for cell in ws[2]:
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Row fills: gold for rank-1, green for positive PnL, red for negative
        for row_idx, r in enumerate(sorted_r, 3):
            pnl_pos = r["metrics"]["total_pnl_dollars"] > 0
            fill = GOLD_FILL if row_idx == 3 else (GREEN_FILL if pnl_pos else RED_FILL)
            for cell in ws[row_idx]:
                cell.fill = fill

        # Column widths
        ws.column_dimensions["A"].width = 6   # Rank
        ws.column_dimensions["B"].width = 38  # Model name
        for i in range(3, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12

        # ── Per-model metrics sheet (one row per model, sorted) ───
        df.to_excel(writer, sheet_name="AllMetrics", index=False)
        ws2 = writer.sheets["AllMetrics"]
        for cell in ws2[1]:
            cell.font = HEADER_FONT



# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main(argv: Optional[List[str]] = None) -> None:
    parser = argparse.ArgumentParser(
        description="Run test-fold evaluation across all saved checkpoints."
    )
    parser.add_argument(
        "--models-dir", required=True,
        help="Path to folder containing checkpoint_*.zip files",
    )
    parser.add_argument(
        "--config", required=True,
        help="Path to config/ directory (YAML files)",
    )
    parser.add_argument(
        "--data", required=True,
        help="Path to data/ directory",
    )
    parser.add_argument(
        "--test-start", default=None,
        help="Override test period start date (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--test-end", default=None,
        help="Override test period end date (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--n-episodes", type=int, default=0,
        help="Episodes per checkpoint. 0 (default) = run every test day exactly once.",
    )
    parser.add_argument(
        "--n-workers", type=int, default=8,
        help="Parallel envs per checkpoint evaluation (default: 8).",
    )
    parser.add_argument(
        "--out-dir", default=None,
        help="Directory to write HTML/Excel journals and results (default: models-dir/test_results)",
    )
    parser.add_argument(
        "--best-only", action="store_true",
        help="Only generate a journal for the single best model (by --rank-by).",
    )
    parser.add_argument(
        "--rank-by",
        choices=["score", "pnl_dollars", "pnl_r", "sharpe"],
        default="score",
        help="Ranking metric for --best-only. Default: score (composite).",
    )
    parser.add_argument(
        "--journal-suffix",
        default="journal",
        help="Suffix appended to journal filenames (<name>_<suffix>.html/.xlsx). Default: journal.",
    )
    args = parser.parse_args(argv)

    models_dir = Path(args.models_dir)
    out_dir = Path(args.out_dir) if args.out_dir else models_dir.parent / "test_results"
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Clean previous results ────────────────────────────────────────────────
    suffix = args.journal_suffix
    removed = 0
    for pattern in ("*.xlsx", f"*_{suffix}.html"):
        for f in out_dir.glob(pattern):
            f.unlink()
            removed += 1
    if removed:
        print(f"  Cleaned {removed} file(s) from {out_dir}\n")

    # ── Build environment components ─────────────────────────────────────────
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

    # Import here to avoid circular imports at module level
    import yaml
    from utils.logger import configure_logging, get_logger as _get_logger
    configure_logging()
    _log = _get_logger(__name__)

    def _load_configs(config_dir: str) -> dict:
        config_dir = Path(config_dir)
        file_map = {
            "agent":       "agent_config.yaml",
            "environment": "environment_config.yaml",
            "features":    "features_config.yaml",
            "risk":        "risk_config.yaml",
            "reward":      "reward_config.yaml",
            "logging":     "logging_config.yaml",
        }
        cfgs = {}
        for key, fname in file_map.items():
            p = config_dir / fname
            cfgs[key] = yaml.safe_load(p.read_text()) if p.exists() else {}
        return cfgs

    configs = _load_configs(args.config)
    env_cfg     = configs["environment"]
    feat_cfg    = configs["features"]
    risk_cfg    = configs["risk"]
    reward_cfg  = configs["reward"]
    agent_cfg   = configs["agent"]

    from data.data_loader     import DataLoader
    from data.data_splitter   import DataSplitter
    from environment.action_space       import ActionMasker
    from environment.position_manager   import PositionManager
    from environment.reward_calculator  import RewardCalculator
    from environment.trading_env        import TradingEnv
    from features.atr_calculator        import ATRCalculator
    from features.harmonic_detector     import HarmonicDetector
    from features.observation_builder   import ObservationBuilder
    from features.order_zone_engine     import OrderZoneEngine
    from features.zone_detector         import ZoneDetector
    from utils.instrument                import load_instrument_profile

    instrument  = env_cfg.get("instruments", {}).get("default", "NQ")
    instrument_profile = load_instrument_profile(env_cfg)
    session_cfg = env_cfg.get("session", {})
    bar_minutes = int(session_cfg.get("bar_timeframe_minutes", 5))

    data_loader = DataLoader(
        data_dir=args.data,
        instrument=instrument,
        intraday_tf=f"{bar_minutes}min",
        daily_tf=session_cfg.get("daily_timeframe", "1D"),
        tz=session_cfg.get("timezone", "America/New_York"),
    )
    data_loader.load()

    all_days = data_loader.get_trading_days()
    atr_cfg  = feat_cfg.get("atr", {})
    atr_calc = ATRCalculator(
        atr_period=atr_cfg.get("period", 14),
        exhaustion_threshold=atr_cfg.get("exhaustion_threshold", 0.95),
    )
    atr_calc.fit(data_loader.daily)

    import pandas as _pd
    trading_days = [
        d for d in all_days
        if _pd.Timestamp(d).weekday() < 5
        and atr_calc.get_atr_for_date(d) is not None
    ]

    # Determine test days
    if args.test_start or args.test_end:
        ts = args.test_start or "1970-01-01"
        te = args.test_end   or "2099-12-31"
        test_days = [d for d in trading_days if ts <= str(d) <= te]
    else:
        split = DataSplitter.split_by_counts(trading_days, n_train=252, n_val=26)
        test_days = split.test

    if not test_days:
        print("ERROR: No test days found — check --test-start / --test-end or data range.")
        sys.exit(1)

    _log.info("Test period", days=len(test_days),
              start=str(test_days[0]), end=str(test_days[-1]))

    # ── Build component factories ─────────────────────────────────────────────
    zones_cfg   = feat_cfg.get("zones", {})
    oz_cfg      = feat_cfg.get("order_zone", {})
    obs_cfg     = env_cfg.get("observation", {})
    atr_gate    = risk_cfg.get("atr_gate", {})
    daily_lim   = risk_cfg.get("daily_limits", {})
    session_risk = risk_cfg.get("session", {})
    account_cfg = env_cfg.get("account", {})
    sizing_cfg  = risk_cfg.get("sizing", {})
    trail_cfg   = risk_cfg.get("trailing", {})
    real_capital = float(account_cfg.get("initial_balance", 100000.0))
    point_value  = instrument_profile.point_value

    harm_cfg = feat_cfg.get("harmonic", {})
    harmonic_detector = HarmonicDetector(
        lookback_bars=harm_cfg.get("lookback_bars", 40),
        pivot_window=harm_cfg.get("pivot_window", 3),
        symmetry_tol_atr_pct=harm_cfg.get("symmetry_tol_atr_pct", 0.12),
        min_peak_atr_pct=harm_cfg.get("min_peak_atr_pct", 0.15),
        min_separation_bars=harm_cfg.get("min_separation_bars", 5),
        recency_bars=harm_cfg.get("recency_bars", 6),
    ) if harm_cfg.get("enabled", True) else None
    observation_builder = ObservationBuilder(
        clip_value=obs_cfg.get("clip_observations", 10.0),
        normalize_observations=obs_cfg.get("normalize_observations", True),
        lookback_bars=obs_cfg.get("lookback_bars", 20),
        max_zone_age_bars=zones_cfg.get("max_zone_age_bars", 300),
        max_zone_pts=instrument_profile.max_zone_pts,
        min_zone_pts=instrument_profile.min_zone_pts,
    )
    order_zone_engine = OrderZoneEngine(
        weights=oz_cfg.get("weights"),
        min_confluence_score=oz_cfg.get("min_confluence_score", 0.60),
        min_rr_ratio=risk_cfg.get("take_profit", {}).get("min_rr_ratio", 1.5),
        max_zone_pts=instrument_profile.max_zone_pts,
        stop_buffer_pts=instrument_profile.stop_buffer_pts,
        fallback_stop_pts=instrument_profile.fallback_stop_pts,
    )
    action_masker = ActionMasker(
        atr_exhaustion_threshold=atr_gate.get("block_entries_above_pct", 0.95),
        trail_min_r=trail_cfg.get("activate_at_r", 2.0),
        max_trades_per_day=daily_lim.get("max_trades_per_day", 5),
        no_entry_last_n_bars=session_risk.get("no_entry_last_n_bars", 3),
        min_bars_between_trades=session_risk.get("min_bars_between_trades", 3),
    )
    action_masker.max_pending_order_bars = session_risk.get("max_pending_order_bars", 5)
    reward_calculator = RewardCalculator.from_config(reward_cfg)

    zone_detector_defaults = {
        "consolidation_min_bars":       2,
        "consolidation_max_bars":       8,
        "consolidation_range_atr_pct":  0.20,
        "impulse_min_body_atr_pct":     0.15,
        "max_zone_age_bars":            200,
        "max_zone_touches":             3,
        "zone_buffer_atr_pct":          0.02,
    }

    def make_position_manager():
        return PositionManager(
            real_capital=real_capital,
            risk_per_trade_pct=sizing_cfg.get("risk_per_trade_pct", 0.01),
            min_contracts=sizing_cfg.get("min_contracts", 1),
            max_contracts=sizing_cfg.get("max_contracts", 2),
            point_value=point_value,
            max_trades_per_day=daily_lim.get("max_trades_per_day", 5),
            trail_activate_r=trail_cfg.get("activate_at_r", 2.0),
            trail_aggressive_r=trail_cfg.get("trail_aggressively_at_r", 4.0),
            trail_lock_in_r=trail_cfg.get("lock_in_r_at_trail", 2.0),
            max_daily_loss_r=daily_lim.get("max_daily_loss_r", 3.0),
            max_daily_loss_dollars=daily_lim.get("max_daily_loss_dollars", 3000.0),
            max_drawdown_r=risk_cfg.get("position", {}).get("max_drawdown_r", 5.0),
            pause_bars_after_loss_streak=daily_lim.get("pause_bars_after_loss_streak", 6),
            loss_streak_threshold=daily_lim.get("max_consecutive_losses_before_pause", 3),
            zone_buffer_atr_pct=risk_cfg.get("stop_loss", {}).get("zone_buffer_atr_pct", 0.03),
            contract_tiers=instrument_profile.contract_tiers,
            confluence_tier_thresholds=instrument_profile.confluence_tier_thresholds,
            max_zone_pts=instrument_profile.max_zone_pts,
        )

    session_start = session_cfg.get("rth_start_utc", "08:30")
    session_end   = session_cfg.get("rth_end_utc",   "15:00")
    session_type  = session_cfg.get("session_type", "RTH").upper()

    def env_factory():
        return TradingEnv(
            data_loader=data_loader,
            trading_days=test_days,
            position_manager=make_position_manager(),
            reward_calculator=reward_calculator,
            observation_builder=observation_builder,
            atr_calculator=atr_calc,
            zone_detector=ZoneDetector(
                **{k: zones_cfg.get(k, v) for k, v in zone_detector_defaults.items()},
                break_buffer_pts=instrument_profile.stop_buffer_pts,
            ),
            order_zone_engine=order_zone_engine,
            action_masker=action_masker,
            instrument=instrument_profile,
            rth_start=session_start,
            rth_end=session_end,
            no_entry_last_n_bars=session_risk.get("no_entry_last_n_bars", 3),
            early_terminate_on_max_dd=env_cfg.get("episode", {}).get(
                "early_termination_on_max_drawdown", True),
            bar_minutes=bar_minutes,
            curriculum_filter_fn=None,
            augmentor=None,
            harmonic_detector=harmonic_detector,
            session_type=session_type,
            random_start=False,
            seed=agent_cfg.get("seed", 42) + 999,
            zone_lookback_bars=feat_cfg.get("zone_lookback_bars", 500),
        )

    # ── Resolve episode count ──────────────────────────────────────────────────
    n_episodes = args.n_episodes if args.n_episodes > 0 else len(test_days)
    _log.info("Episodes per checkpoint", n_episodes=n_episodes,
              test_days=len(test_days))

    # ── Find and evaluate checkpoints ────────────────────────────────────────
    checkpoints = _find_checkpoints(models_dir)
    n_ckpt = sum(1 for c in checkpoints if "checkpoint_" in c.name)
    n_hot  = sum(1 for c in checkpoints if "hotsave_"    in c.name)
    print(f"\nFound {len(checkpoints)} model(s) in {models_dir}  "
          f"({n_ckpt} checkpoints, {n_hot} hotsaves)\n"
          f"Test period : {test_days[0]} → {test_days[-1]}  ({len(test_days)} trading days)\n"
          f"Episodes    : {n_episodes} per model\n")

    results: List[Dict] = []
    sep = _separator()

    print(sep)
    print(_header_row())
    print(sep)

    for ckpt in checkpoints:
        name = _checkpoint_base(ckpt)
        vn   = _vecnorm_path(ckpt)

        print(f"  → Evaluating {name} ...", end="", flush=True)
        try:
            trades, episodes = _run_checkpoint(
                ckpt, vn, env_factory,
                n_episodes=n_episodes,
                n_workers=args.n_workers,
            )
        except Exception as exc:
            print(f" FAILED: {exc}")
            continue

        # Enrich trades with entry/exit timestamps
        try:
            trades = _resolve_trade_times(trades, data_loader)
        except Exception:
            pass

        m     = _compute_metrics(trades, episodes)
        score = _composite(m)

        print(f"\r{_data_row(name, m, score)}")

        results.append({
            "name":      name,
            "ckpt_path": ckpt,
            "metrics":   m,
            "score":     score,
            "trades":    trades,
            "episodes":  episodes,
        })

    print(sep)

    if not results:
        print("No results to rank.")
        return

    # ── Ranked summary (ascending by rank metric — best last) ────────────────
    _rank_key_fns = {
        "score":        lambda r: r["score"],
        "pnl_dollars":  lambda r: r["metrics"]["total_pnl_dollars"],
        "pnl_r":        lambda r: r["metrics"]["total_pnl_r"],
        "sharpe":       lambda r: r["metrics"]["sharpe"],
    }
    rank_fn = _rank_key_fns[args.rank_by]
    results.sort(key=rank_fn)

    print(f"\n── Ranked Summary (by {args.rank_by}, best last) ──────────────────────\n")
    print(sep)
    print(_header_row())
    print(sep)
    for r in results:
        print(_data_row(r["name"], r["metrics"], r["score"]))
    print(sep)

    best = results[-1]
    bm = best["metrics"]
    print(f"\n  Best checkpoint (by {args.rank_by}): {best['name']}")
    print(f"  Score={best['score']:.3f}  "
          f"Trades={bm['n_trades']}  "
          f"WR={bm['win_rate']*100:.1f}%  "
          f"PnL=${bm['total_pnl_dollars']:+,.0f}  "
          f"$PF={bm['profit_factor_usd']:.2f}  "
          f"Sharpe={bm['sharpe']:.3f}")

    # ── Write per-model journals (best-only or all) ───────────────────────────
    to_journal = [best] if args.best_only else results
    print(f"\n  Writing {len(to_journal)} journal file(s) to {out_dir} ...")
    for r in to_journal:
        name_r = r["name"]
        if PLOTLY_OK:
            journal_path = out_dir / f"{name_r}_{suffix}.html"
            try:
                _build_journal(
                    name_r, r["trades"], r["metrics"], r["score"],
                    data_loader, test_days, journal_path,
                )
            except Exception as exc:
                _log.warning("Journal HTML failed", ckpt=name_r, error=str(exc))

        excel_path = out_dir / f"{name_r}_{suffix}.xlsx"
        try:
            _build_excel_journal(
                name_r, r["trades"], r["metrics"], r["score"],
                test_days, excel_path,
            )
        except Exception as exc:
            _log.warning("Journal Excel failed", ckpt=name_r, error=str(exc))

    # ── Leaderboard Excel (all models, ranked) ────────────────────────────────
    lb_path = out_dir / "leaderboard.xlsx"
    try:
        _build_leaderboard_excel(results, test_days, lb_path)
    except Exception as exc:
        _log.warning("Leaderboard Excel failed", error=str(exc))

    # ── Write text results ────────────────────────────────────────────────────
    txt_path = out_dir / "test_fold_results.txt"
    lines: List[str] = [
        f"Test Fold Results",
        f"Models dir : {models_dir}",
        f"Test period: {test_days[0]} → {test_days[-1]}  ({len(test_days)} days)",
        f"Episodes   : {n_episodes} per checkpoint",
        "",
        sep,
        _header_row(),
        sep,
    ]
    for r in sorted(results, key=lambda r: r["score"], reverse=True):
        lines.append(_data_row(r["name"], r["metrics"], r["score"]))
    lines += [sep, "", f"Best: {best['name']}  score={best['score']:.4f}"]
    txt_path.write_text("\n".join(lines), encoding="utf-8")

    htmls   = list(out_dir.glob(f"*_{suffix}.html"))
    excels  = list(out_dir.glob(f"*_{suffix}.xlsx"))
    print(f"\n  Leaderboard  → {lb_path}")
    print(f"  Results txt  → {txt_path}")
    print(f"  HTML journals: {len(htmls)} file(s)")
    print(f"  Excel journals: {len(excels)} file(s)")
    print(f"\n  Output dir → {out_dir}")


if __name__ == "__main__":
    main()
