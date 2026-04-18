"""
training/training_journal_callback.py
=======================================
SB3 callback that accumulates every trade from all training envs and
periodically writes:

  logs/journal/training_journal.xlsx   — Excel workbook (3 sheets)
  logs/journal/training_journal.html   — Interactive Plotly chart

Both files are overwritten on each save so you always have the latest
snapshot without accumulating multiple files.

Excel sheets
------------
  Trades   : one row per trade — entry/SL/TP/close, PnL, duration, etc.
  Daily    : per-date aggregates (trades, win rate, total PnL)
  Summary  : overall training stats (win rate, Sharpe, profit factor …)

Plotly chart
------------
  Row 1: Cumulative equity curve (R)
  Row 2: Per-trade PnL bar chart (green=win, red=loss)
  Row 3: Cumulative drawdown (R)
  Row 4: Trade duration histogram
  Bottom: summary stats table
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, List

import numpy as np
from stable_baselines3.common.callbacks import BaseCallback


# ── Colour constants (dark theme) ────────────────────────────────────────────
_GREEN = "#26a69a"
_RED   = "#ef5350"
_BLUE  = "#42a5f5"
_GOLD  = "#ffd700"
_BG    = "#131722"
_PAPER = "#1e2230"
_GRID  = "#2a2e39"
_TEXT  = "#d1d4dc"


class TrainingJournalCallback(BaseCallback):
    """
    Collects trades from all envs during training and saves Excel + Plotly.

    Parameters
    ----------
    journal_dir : str | Path
        Output directory (created if missing).
    save_every_steps : int
        Write files after every N timesteps (default 50 000).
    data_dir : str | Path | None
        If provided, also generates a comprehensive OHLC trade chart HTML
        alongside the regular equity-curve HTML.
    instrument : str
        Instrument ticker matching the CSV in data_dir (default "ES").
    bar_minutes : int
        Bar timeframe in minutes (must match the CSV data).
    verbose : int
    """

    def __init__(
        self,
        journal_dir: str | Path = "logs/journal",
        save_every_steps: int = 50_000,
        data_dir: str | Path | None = None,
        instrument: str = "ES",
        bar_minutes: int = 5,
        verbose: int = 0,
    ) -> None:
        super().__init__(verbose)
        self.journal_dir      = Path(journal_dir)
        self.save_every_steps = save_every_steps
        self.data_dir         = Path(data_dir) if data_dir else None
        self.instrument       = instrument
        self.bar_minutes      = bar_minutes
        self._trades: List[dict] = []
        self._last_save_step  = 0
        self._save_n          = 0   # snapshot counter — incremented on every save

    # ── SB3 hooks ─────────────────────────────────────────────────────────────

    def _on_step(self) -> bool:
        infos = self.locals.get("infos", [])
        dones = self.locals.get("dones", [])
        for env_idx, (info, done) in enumerate(zip(infos, dones)):
            if done:
                for t in info.get("trades_list", []):
                    t["global_step"] = self.num_timesteps
                    t["env_id"]      = env_idx
                    self._trades.append(t)

        if (self.num_timesteps - self._last_save_step) >= self.save_every_steps:
            self._save()
            self._last_save_step = self.num_timesteps

        return True

    def _on_training_end(self) -> None:
        if self._trades:
            self._save()

    # ── Save ──────────────────────────────────────────────────────────────────

    def _save(self) -> None:
        if not self._trades:
            return
        self._save_n += 1
        snap_dir = self.journal_dir / "snapshots"
        snap_dir.mkdir(parents=True, exist_ok=True)
        self.journal_dir.mkdir(parents=True, exist_ok=True)
        stem = f"journal_s{self._save_n:04d}_step{self.num_timesteps:010d}"
        try:
            self._write_excel(snap_dir, stem=stem)
        except Exception as exc:
            if self.verbose:
                print(f"[TrainingJournal] Excel write failed: {exc}")
        try:
            self._write_plotly(snap_dir, stem=stem)
        except Exception as exc:
            if self.verbose:
                print(f"[TrainingJournal] Plotly write failed: {exc}")
        self._write_trade_chart(snap_dir, stem=stem)

    # ── Public snapshot API (called by hotsave callback) ─────────────────────

    def write_snapshot(self, output_dir: Path, stem: str, trades: list | None = None) -> None:
        """Write Excel + HTML to output_dir/<stem>.{xlsx,html} without touching 'latest'.

        trades : if provided, use this list instead of self._trades (used by hotsave
                 callback which maintains its own independent accumulator).
        """
        _trades = trades if trades is not None else self._trades
        if not _trades:
            return
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        try:
            self._write_excel(output_dir, stem=stem, copy_latest=False, trades=_trades)
        except Exception as exc:
            if self.verbose:
                print(f"[TrainingJournal] Hotsave Excel write failed: {exc}")
        try:
            self._write_plotly(output_dir, stem=stem, copy_latest=False, trades=_trades)
        except Exception as exc:
            if self.verbose:
                print(f"[TrainingJournal] Hotsave Plotly write failed: {exc}")
        self._write_trade_chart(output_dir, stem=stem, trades=_trades)

    # ── Trade chart (OHLC + annotations) ─────────────────────────────────────

    def _write_trade_chart(
        self,
        out_dir: Path,
        stem:    str,
        trades:  list | None = None,
    ) -> None:
        if self.data_dir is None:
            return
        _trades = trades if trades is not None else self._trades
        if not _trades:
            return
        import shutil
        from training.trade_chart import write_trade_chart

        # ── Aggregate OHLC chart (all envs combined) ──────────────────────────
        try:
            out_path = out_dir / f"{stem}_trades.html"
            write_trade_chart(
                trades=_trades,
                data_dir=self.data_dir,
                output_path=out_path,
                instrument=self.instrument,
                bar_minutes=self.bar_minutes,
                title_prefix=f"All Envs — step {self.num_timesteps:,}",
            )
            shutil.copy2(str(out_path), str(self.journal_dir / "training_trades.html"))
            if self.verbose:
                print(f"[TrainingJournal] Aggregate trade chart → {out_path}")
        except Exception as exc:
            if self.verbose:
                print(f"[TrainingJournal] Aggregate trade chart failed: {exc}")

        # ── One OHLC chart per env ────────────────────────────────────────────
        env_ids = sorted(set(t.get("env_id", 0) for t in _trades))
        if len(env_ids) <= 1:
            return   # single env — aggregate chart is sufficient
        for eid in env_ids:
            env_trades = [t for t in _trades if t.get("env_id", 0) == eid]
            if not env_trades:
                continue
            try:
                env_path = out_dir / f"{stem}_env{eid:02d}.html"
                write_trade_chart(
                    trades=env_trades,
                    data_dir=self.data_dir,
                    output_path=env_path,
                    instrument=self.instrument,
                    bar_minutes=self.bar_minutes,
                    title_prefix=f"Env {eid:02d} — step {self.num_timesteps:,}",
                )
                # always keep latest per-env copy
                shutil.copy2(
                    str(env_path),
                    str(self.journal_dir / f"training_trades_env{eid:02d}.html"),
                )
                if self.verbose:
                    print(f"[TrainingJournal] Env {eid:02d} trade chart → {env_path}")
            except Exception as exc:
                if self.verbose:
                    print(f"[TrainingJournal] Env {eid:02d} trade chart failed: {exc}")

    # ── Excel ─────────────────────────────────────────────────────────────────

    def _write_excel(self, snap_dir: Path, stem: str | None = None, copy_latest: bool = True, trades: list | None = None) -> None:
        import pandas as pd
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        df = pd.DataFrame(trades if trades is not None else self._trades)
        if stem is None:
            stem = f"journal_s{self._save_n:04d}_step{self.num_timesteps:010d}"
        path = snap_dir / f"{stem}.xlsx"

        with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
            # ── Sheet 1: Trades ───────────────────────────────
            cols = [
                "global_step", "date", "direction",
                "entry_price", "stop_price", "initial_target", "exit_price",
                "pnl_r", "pnl_dollars", "pnl_points",
                "n_contracts", "duration_min", "exit_reason",
                "is_win", "is_rth", "mae_r",
            ]
            trades_df = df[[c for c in cols if c in df.columns]].copy()
            trades_df.to_excel(writer, sheet_name="Trades", index=False)
            ws = writer.sheets["Trades"]
            _style_trades_sheet(ws, trades_df)

            # ── Sheet 2: Daily ────────────────────────────────
            if "date" in df.columns:
                def _daily_row(g):
                    n_g   = len(g)
                    wins_g  = g[g["is_win"]]
                    losses_g = g[~g["is_win"]]
                    pnl_r_g  = float(g["pnl_r"].sum())
                    pnl_d_g  = float(g["pnl_dollars"].sum()) if "pnl_dollars" in g.columns else 0.0
                    row = {
                        "trades":            n_g,
                        "wins":              len(wins_g),
                        "total_pnl_r":       round(pnl_r_g, 4),
                        "total_pnl_dollars": round(pnl_d_g, 2),
                        "avg_pnl_r":         round(pnl_r_g / n_g, 4) if n_g else 0.0,
                        "avg_duration_min":  round(float(g["duration_min"].mean()), 1) if "duration_min" in g.columns else 0.0,
                        "win_rate":          round(len(wins_g) / n_g, 3) if n_g else 0.0,
                    }
                    if "is_rth" in g.columns:
                        rth_g = g[g["is_rth"]]
                        eth_g = g[~g["is_rth"]]
                        row["rth_pnl_$"] = round(float(rth_g["pnl_dollars"].sum()) if "pnl_dollars" in rth_g.columns and len(rth_g) else 0.0, 2)
                        row["rth_pf"]    = round(_pf(rth_g), 2)
                        row["eth_pnl_$"] = round(float(eth_g["pnl_dollars"].sum()) if "pnl_dollars" in eth_g.columns and len(eth_g) else 0.0, 2)
                        row["eth_pf"]    = round(_pf(eth_g), 2)
                    return pd.Series(row)

                daily = df.groupby("date").apply(_daily_row).reset_index()
                daily.to_excel(writer, sheet_name="Daily", index=False)
                _style_daily_sheet(writer.sheets["Daily"], daily)

            # ── Sheet 3: Summary ──────────────────────────────
            summary = _compute_summary(df, self.num_timesteps)
            summary_df = pd.DataFrame(list(summary.items()), columns=["Metric", "Value"])
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            _style_summary_sheet(writer.sheets["Summary"])

            # ── Sheets 4+: Per-env ────────────────────────────
            if "env_id" in df.columns:
                _TRADE_COLS = [
                    "global_step", "date", "direction",
                    "entry_price", "stop_price", "initial_target", "exit_price",
                    "pnl_r", "pnl_dollars", "pnl_points",
                    "n_contracts", "duration_min", "exit_reason",
                    "is_win", "is_rth", "mae_r",
                ]
                for env_id, env_df in df.groupby("env_id"):
                    sheet = f"Env_{int(env_id):02d}"
                    ecols = [c for c in _TRADE_COLS if c in env_df.columns]
                    env_df[ecols].to_excel(writer, sheet_name=sheet, index=False)
                    _style_trades_sheet(writer.sheets[sheet], env_df[ecols])

        if copy_latest:
            import shutil
            latest = self.journal_dir / "training_journal.xlsx"
            shutil.copy2(str(path), str(latest))

        if self.verbose:
            print(f"[TrainingJournal] Excel saved → {path}  ({len(df)} trades)")

    # ── Plotly ────────────────────────────────────────────────────────────────

    def _write_plotly(self, snap_dir: Path, stem: str | None = None, copy_latest: bool = True, trades: list | None = None) -> None:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots

        df_raw = trades if trades is not None else self._trades
        if not df_raw:
            return

        pnl_r  = [t["pnl_r"]         for t in df_raw]
        pnl_d  = [t.get("pnl_dollars", t["pnl_r"]) for t in df_raw]
        is_win = [t["is_win"]         for t in df_raw]
        durs   = [t["duration_min"]   for t in df_raw]
        dates  = [t["date"]           for t in df_raw]
        steps  = [t["global_step"]    for t in df_raw]
        dirs   = [t["direction"]      for t in df_raw]

        n = len(pnl_r)
        trade_nums = list(range(1, n + 1))
        equity_d   = np.cumsum(pnl_d)   # dollars

        # Cumulative drawdown in dollars: equity − running peak (always ≤ 0)
        peak_d     = np.maximum.accumulate(equity_d)
        drawdown_d = list(equity_d - peak_d)

        total_d   = sum(pnl_d)
        n_wins    = sum(is_win)
        wr_pct    = 100 * n_wins / n if n else 0
        color_r   = _GREEN if total_d >= 0 else _RED

        fig = make_subplots(
            rows=6, cols=1,
            row_heights=[0.22, 0.16, 0.12, 0.12, 0.16, 0.22],
            vertical_spacing=0.035,
            specs=[
                [{"type": "scatter"}],
                [{"type": "bar"}],
                [{"type": "scatter"}],
                [{"type": "histogram"}],
                [{"type": "table"}],
                [{"type": "table"}],
            ],
            subplot_titles=[
                "Cumulative Equity ($)",
                "Per-Trade PnL ($)",
                "Cumulative Drawdown ($)",
                "Trade Duration Distribution (min)",
                "Overall Summary",
                "Per-Environment Breakdown",
            ],
        )

        # ── Row 1: Equity curve ───────────────────────────────
        fig.add_trace(
            go.Scatter(
                x=trade_nums, y=equity_d,
                mode="lines",
                line=dict(color=_BLUE, width=2),
                fill="tozeroy",
                fillcolor="rgba(66,165,245,0.1)",
                hovertemplate="Trade %{x}<br>Equity: $%{y:,.0f}<extra></extra>",
                name="Equity",
            ),
            row=1, col=1,
        )
        fig.add_hline(y=0, line=dict(color=_GRID, width=1), row=1, col=1)

        # ── Row 2: Per-trade PnL bars ─────────────────────────
        bar_colors = [_GREEN if w else _RED for w in is_win]
        fig.add_trace(
            go.Bar(
                x=trade_nums,
                y=pnl_d,
                marker_color=bar_colors,
                hovertemplate="Trade %{x}<br>PnL: $%{y:,.0f}<extra></extra>",
                name="PnL",
            ),
            row=2, col=1,
        )
        fig.add_hline(y=0, line=dict(color=_GRID, width=1), row=2, col=1)

        # ── Row 3: Cumulative drawdown ────────────────────────
        fig.add_trace(
            go.Scatter(
                x=trade_nums, y=drawdown_d,
                mode="lines",
                line=dict(color=_RED, width=1.5),
                fill="tozeroy",
                fillcolor="rgba(239,83,80,0.15)",
                hovertemplate="Trade %{x}<br>Drawdown: $%{y:,.0f}<extra></extra>",
                name="Drawdown",
            ),
            row=3, col=1,
        )
        fig.add_hline(y=0, line=dict(color=_GRID, width=1, dash="dot"), row=3, col=1)

        # ── Row 4: Duration histogram ─────────────────────────
        fig.add_trace(
            go.Histogram(
                x=durs,
                marker_color=_BLUE,
                opacity=0.7,
                hovertemplate="Duration: %{x}m<br>Count: %{y}<extra></extra>",
                name="Duration",
            ),
            row=4, col=1,
        )

        # ── Row 5: Overall summary table ─────────────────────
        import pandas as pd
        summary = _compute_summary(pd.DataFrame(df_raw), self.num_timesteps)
        s_keys = list(summary.keys())
        s_vals = [str(v) for v in summary.values()]
        half   = len(s_keys) // 2 + len(s_keys) % 2

        fig.add_trace(
            go.Table(
                header=dict(
                    values=["<b>Metric</b>", "<b>Value</b>",
                            "<b>Metric</b>", "<b>Value</b>"],
                    fill_color=_PAPER,
                    font=dict(color=_TEXT, size=11),
                    align="center",
                    line_color=_GRID,
                    height=26,
                ),
                cells=dict(
                    values=[
                        s_keys[:half],  s_vals[:half],
                        s_keys[half:],  s_vals[half:],
                    ],
                    fill_color=_BG,
                    font=dict(color=_TEXT, size=10),
                    align=["left", "right", "left", "right"],
                    line_color=_GRID,
                    height=22,
                ),
            ),
            row=5, col=1,
        )

        # ── Row 6: Per-environment breakdown table ────────────
        from training.fold_journal_callback import _env_summary_row
        env_rows = []
        env_ids_seen = sorted(set(t.get("env_id", 0) for t in df_raw))
        env_df_all = pd.DataFrame(df_raw)
        for eid in env_ids_seen:
            env_slice = env_df_all[env_df_all["env_id"] == eid] if "env_id" in env_df_all.columns else env_df_all
            env_rows.append(_env_summary_row(eid, env_slice))
        if env_rows:
            tbl_df = pd.DataFrame(env_rows)
            fig.add_trace(
                go.Table(
                    header=dict(
                        values=[f"<b>{c}</b>" for c in tbl_df.columns],
                        fill_color=_PAPER,
                        font=dict(color=_TEXT, size=10),
                        align="center",
                        line_color=_GRID,
                        height=24,
                    ),
                    cells=dict(
                        values=[tbl_df[c].tolist() for c in tbl_df.columns],
                        fill_color=_BG,
                        font=dict(color=_TEXT, size=10),
                        align=["center"] * len(tbl_df.columns),
                        line_color=_GRID,
                        height=20,
                    ),
                ),
                row=6, col=1,
            )

        # ── Layout ────────────────────────────────────────────
        title = (
            f"Training Journal  —  Step {self.num_timesteps:,}  |  "
            f"{n} trades  |  WR {wr_pct:.0f}%  |  "
            f"<span style='color:{color_r}'>{total_r:+.2f}R</span>"
        )
        fig.update_layout(
            title=dict(text=title, font=dict(size=14, color=_TEXT), x=0.5),
            paper_bgcolor=_PAPER,
            plot_bgcolor=_BG,
            font=dict(color=_TEXT, family="monospace"),
            showlegend=False,
            dragmode="pan",
            margin=dict(l=60, r=30, t=60, b=20),
            height=1500,
        )
        for row in range(1, 5):
            fig.update_xaxes(gridcolor=_GRID, zeroline=False,
                             rangeslider=dict(visible=False),
                             row=row, col=1)
            fig.update_yaxes(gridcolor=_GRID, zeroline=False, row=row, col=1)
        # Rangeslider on bottom data row (row 4) for horizontal scrolling
        fig.update_xaxes(
            rangeslider=dict(visible=True, thickness=0.04, bgcolor=_PAPER),
            row=4, col=1,
        )

        if stem is None:
            stem = f"journal_s{self._save_n:04d}_step{self.num_timesteps:010d}"
        path = snap_dir / f"{stem}.html"
        fig.write_html(
            str(path),
            include_plotlyjs="cdn",
            config={"scrollZoom": True, "displayModeBar": True,
                    "modeBarButtonsToAdd": ["pan2d", "zoom2d"]},
        )

        if copy_latest:
            import shutil
            latest = self.journal_dir / "training_journal.html"
            shutil.copy2(str(path), str(latest))

        if self.verbose:
            print(f"[TrainingJournal] Chart  saved → {path}")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _pf(df: Any) -> float:
    """Profit factor for a trades DataFrame slice."""
    if df is None or len(df) == 0:
        return 0.0
    wins   = df[df["is_win"]] if "is_win" in df.columns else df.iloc[:0]
    losses = df[~df["is_win"]] if "is_win" in df.columns else df.iloc[:0]
    gw = float(wins["pnl_r"].sum())           if len(wins)   else 0.0
    gl = abs(float(losses["pnl_r"].sum()))    if len(losses) else 1e-6
    return min(gw / max(gl, 1e-6), 99.99)


def _compute_summary(df: Any, step: int) -> dict:
    """Compute summary stats from a trades DataFrame."""
    import pandas as pd
    if df.empty:
        return {"error": "no trades"}

    n      = len(df)
    wins   = df[df["is_win"]]
    losses = df[~df["is_win"]]
    wr     = len(wins) / n if n else 0.0
    total_r  = float(df["pnl_r"].sum())
    total_usd = float(df["pnl_dollars"].sum()) if "pnl_dollars" in df else 0.0
    avg_win  = float(wins["pnl_r"].mean())  if len(wins)   else 0.0
    avg_loss = float(losses["pnl_r"].mean()) if len(losses) else 0.0
    gross_w  = float(wins["pnl_r"].sum())   if len(wins)   else 0.0
    gross_l  = abs(float(losses["pnl_r"].sum())) if len(losses) else 1e-6
    pf_val   = min(gross_w / max(gross_l, 1e-6), 99.99)
    avg_dur  = float(df["duration_min"].mean()) if "duration_min" in df else 0.0

    tr = df["pnl_r"].values
    if len(tr) >= 5 and tr.std() > 0.01:
        sharpe = float(np.clip(tr.mean() / tr.std(), -9.99, 9.99))
    else:
        sharpe = 0.0

    # RTH / ETH profit factors
    if "is_rth" in df.columns:
        rth_df   = df[df["is_rth"]]
        eth_df   = df[~df["is_rth"]]
        rth_pnl  = float(rth_df["pnl_dollars"].sum()) if "pnl_dollars" in rth_df.columns and len(rth_df) else 0.0
        eth_pnl  = float(eth_df["pnl_dollars"].sum()) if "pnl_dollars" in eth_df.columns and len(eth_df) else 0.0
        rth_pf_v = _pf(rth_df)
        eth_pf_v = _pf(eth_df)
    else:
        rth_pnl = eth_pnl = rth_pf_v = eth_pf_v = 0.0

    # Max drawdown in dollars
    if "pnl_dollars" in df.columns and len(df):
        eq_d   = np.cumsum(df["pnl_dollars"].values.astype(float))
        peak_d = np.maximum.accumulate(eq_d)
        max_dd = float((peak_d - eq_d).max())
        dd_pct = max_dd / max(float(peak_d.max()), 1.0) * 100 if peak_d.max() > 0 else 0.0
    else:
        max_dd = dd_pct = 0.0

    return {
        "Training Step":    f"{step:,}",
        "Total Trades":     n,
        "Win Rate":         f"{wr*100:.1f}%",
        "Total PnL (R)":    f"{total_r:+.2f}",
        "Total PnL ($)":    f"${total_usd:+,.0f}",
        "Profit Factor":    f"{pf_val:.2f}",
        "RTH PnL ($)":      f"${rth_pnl:+,.0f}",
        "RTH PF":           f"{rth_pf_v:.2f}",
        "ETH PnL ($)":      f"${eth_pnl:+,.0f}",
        "ETH PF":           f"{eth_pf_v:.2f}",
        "Sharpe Ratio":     f"{sharpe:.2f}",
        "Avg Win (R)":      f"{avg_win:+.3f}",
        "Avg Loss (R)":     f"{avg_loss:+.3f}",
        "Avg Duration":     f"{avg_dur:.0f} min",
        "DD%":              f"{dd_pct:.1f}%",
        "Max DD ($)":       f"${max_dd:,.0f}",
        "Total Wins":       len(wins),
        "Total Losses":     len(losses),
    }


def _style_trades_sheet(ws: Any, df: Any) -> None:
    """Apply green/red row fills and auto column widths to Trades sheet."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    green_fill = PatternFill("solid", fgColor="C8E6C9")
    red_fill   = PatternFill("solid", fgColor="FFCDD2")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    is_win_col = None
    for i, cell in enumerate(ws[1], 1):
        if cell.value == "is_win":
            is_win_col = i
            break

    for row in ws.iter_rows(min_row=2):
        if is_win_col:
            win_val = row[is_win_col - 1].value
            fill = green_fill if win_val else red_fill
            for cell in row:
                cell.fill = fill

    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 20)


def _style_daily_sheet(ws: Any, df: Any) -> None:
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    green_fill = PatternFill("solid", fgColor="C8E6C9")
    red_fill   = PatternFill("solid", fgColor="FFCDD2")

    for cell in ws[1]:
        cell.font = Font(bold=True)

    pnl_col = None
    for i, cell in enumerate(ws[1], 1):
        if cell.value == "total_pnl_r":
            pnl_col = i
            break

    for row in ws.iter_rows(min_row=2):
        if pnl_col:
            val = row[pnl_col - 1].value
            fill = green_fill if (val or 0) >= 0 else red_fill
            for cell in row:
                cell.fill = fill

    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 18)


def _style_summary_sheet(ws: Any) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
