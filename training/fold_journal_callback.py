"""
training/fold_journal_callback.py
===================================
SB3 callback that tracks trades per training environment during a single
walk-forward fold.  At the end of the fold the caller invokes ``save()``
which writes:

  <fold_dir>/fold_<N>_training_journal.xlsx
      — one sheet per env  (Env_00, Env_01, …)
      — one aggregate Summary sheet

  <fold_dir>/fold_<N>_training_journal.html
      — equity curves for every env overlaid on one chart
      — per-env PnL bar chart
      — cumulative drawdown (R)
      — aggregate stats table

Usage
-----
    cb = FoldJournalCallback(n_envs=16, verbose=1)
    trainer.run()   # or agent.train(..., callback=[..., cb])
    cb.save(fold_id=0, fold_dir=Path("logs/walk_forward/fold_00"))
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Dict, List

import numpy as np
from stable_baselines3.common.callbacks import BaseCallback


# ── Colour palette (dark theme, matches training_journal_callback) ────────────
_COLOURS = [
    "#42a5f5",  # blue
    "#26a69a",  # teal
    "#ffd700",  # gold
    "#ef5350",  # red
    "#ab47bc",  # purple
    "#ff7043",  # orange
    "#66bb6a",  # green
    "#ec407a",  # pink
    "#26c6da",  # cyan
    "#d4e157",  # lime
    "#8d6e63",  # brown
    "#78909c",  # blue-grey
    "#ff8a65",  # deep orange light
    "#ba68c8",  # light purple
    "#4db6ac",  # teal light
    "#aed581",  # light green
]
_GREEN = "#26a69a"
_RED   = "#ef5350"
_BG    = "#131722"
_PAPER = "#1e2230"
_GRID  = "#2a2e39"
_TEXT  = "#d1d4dc"


class FoldJournalCallback(BaseCallback):
    """
    Collects per-environment trades during a walk-forward fold.

    Parameters
    ----------
    n_envs : int
        Number of parallel training environments (for pre-allocating dicts).
    verbose : int
    """

    def __init__(self, n_envs: int = 16, verbose: int = 0) -> None:
        super().__init__(verbose)
        self.n_envs = n_envs
        # env_idx → list of trade dicts
        self._env_trades: Dict[int, List[dict]] = defaultdict(list)

    # ── SB3 hooks ─────────────────────────────────────────────────────────────

    def _on_step(self) -> bool:
        infos = self.locals.get("infos", [])
        dones = self.locals.get("dones", [])
        for env_idx, (info, done) in enumerate(zip(infos, dones)):
            if done:
                for t in info.get("trades_list", []):
                    t_copy = dict(t)
                    t_copy["global_step"] = self.num_timesteps
                    t_copy["env_id"]      = env_idx
                    self._env_trades[env_idx].append(t_copy)
        return True

    # ── Public API ────────────────────────────────────────────────────────────

    def save(self, fold_id: int, fold_dir: Path | str) -> None:
        """
        Write Excel + HTML for this fold.

        Parameters
        ----------
        fold_id : int
            Zero-based fold index (used in filenames and titles).
        fold_dir : Path | str
            Output directory.  Created if it does not exist.
        """
        fold_dir = Path(fold_dir)
        fold_dir.mkdir(parents=True, exist_ok=True)

        all_trades: List[dict] = []
        for trades in self._env_trades.values():
            all_trades.extend(trades)

        if not all_trades:
            if self.verbose:
                print(f"[FoldJournal] Fold {fold_id}: no trades to save.")
            return

        try:
            self._write_excel(fold_id, fold_dir, all_trades)
        except Exception as exc:
            if self.verbose:
                print(f"[FoldJournal] Excel write failed: {exc}")

        try:
            self._write_html(fold_id, fold_dir, all_trades)
        except Exception as exc:
            if self.verbose:
                print(f"[FoldJournal] HTML write failed: {exc}")

    # ── Excel ─────────────────────────────────────────────────────────────────

    def _write_excel(
        self,
        fold_id: int,
        fold_dir: Path,
        all_trades: List[dict],
    ) -> None:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        path = fold_dir / f"fold_{fold_id:02d}_training_journal.xlsx"

        _TRADE_COLS = [
            "global_step", "env_id", "date", "direction",
            "entry_price", "stop_price", "initial_target", "exit_price",
            "pnl_r", "pnl_dollars", "pnl_points",
            "n_contracts", "duration_min", "exit_reason",
            "is_win", "mae_r",
        ]

        with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
            # ── One sheet per env ──────────────────────────────
            env_ids = sorted(self._env_trades.keys())
            for env_idx in env_ids:
                trades = self._env_trades[env_idx]
                if not trades:
                    continue
                df = pd.DataFrame(trades)
                sheet_name = f"Env_{env_idx:02d}"
                cols = [c for c in _TRADE_COLS if c in df.columns]
                df[cols].to_excel(writer, sheet_name=sheet_name, index=False)
                _style_trades_sheet(writer.sheets[sheet_name], df[cols])

            # ── Summary sheet ─────────────────────────────────
            all_df = pd.DataFrame(all_trades)
            per_env_summaries = []
            for env_idx in env_ids:
                env_df = pd.DataFrame(self._env_trades.get(env_idx, []))
                row = _env_summary_row(env_idx, env_df)
                per_env_summaries.append(row)

            summary_df = pd.DataFrame(per_env_summaries)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            _style_summary_sheet(writer.sheets["Summary"], summary_df)

        n_total = len(all_trades)
        if self.verbose:
            print(f"[FoldJournal] Excel saved → {path}  ({n_total} trades across {len(env_ids)} envs)")

    # ── HTML (Plotly) ─────────────────────────────────────────────────────────

    def _write_html(
        self,
        fold_id: int,
        fold_dir: Path,
        all_trades: List[dict],
    ) -> None:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots

        env_ids = sorted(self._env_trades.keys())
        if not env_ids:
            return

        # Sort all trades by global_step then env_id for a clean chronological order
        sorted_trades = sorted(
            all_trades,
            key=lambda t: (t.get("global_step", 0), t.get("env_id", 0)),
        )

        agg_pnl    = [t["pnl_r"]  for t in sorted_trades]
        agg_is_win = [t["is_win"] for t in sorted_trades]
        n_agg      = len(agg_pnl)
        trade_nums = list(range(1, n_agg + 1))

        # ── Build figure: 4 rows ──────────────────────────────
        # Row 1: single aggregate cumulative PnL line
        # Row 2: per-trade PnL bars
        # Row 3: cumulative drawdown
        # Row 4: per-env summary table
        fig = make_subplots(
            rows=4, cols=1,
            row_heights=[0.30, 0.22, 0.18, 0.30],
            vertical_spacing=0.05,
            specs=[
                [{"type": "scatter"}],
                [{"type": "bar"}],
                [{"type": "scatter"}],
                [{"type": "table"}],
            ],
            subplot_titles=[
                "Cumulative PnL (R) — all envs combined",
                "Per-Trade PnL (R)",
                "Cumulative Drawdown (R)",
                "",
            ],
        )

        # ── Row 1: single cumulative PnL line ─────────────────
        equity   = list(np.cumsum(agg_pnl))
        eq_color = _GREEN if (equity[-1] if equity else 0) >= 0 else _RED
        fig.add_trace(
            go.Scatter(
                x=trade_nums,
                y=equity,
                mode="lines",
                line=dict(color=eq_color, width=2),
                fill="tozeroy",
                fillcolor=f"rgba(38,166,154,0.10)" if eq_color == _GREEN
                          else "rgba(239,83,80,0.10)",
                name="Cumulative PnL",
                hovertemplate="Trade #%{x}<br>Cumulative PnL: %{y:+.2f}R<extra></extra>",
            ),
            row=1, col=1,
        )
        fig.add_hline(y=0, line=dict(color=_GRID, width=1), row=1, col=1)

        # ── Row 2: per-trade PnL bars ──────────────────────────
        bar_colors = [_GREEN if w else _RED for w in agg_is_win]
        fig.add_trace(
            go.Bar(
                x=trade_nums,
                y=agg_pnl,
                marker_color=bar_colors,
                hovertemplate="Trade #%{x}<br>PnL: %{y:+.2f}R<extra></extra>",
                name="PnL",
                showlegend=False,
            ),
            row=2, col=1,
        )
        fig.add_hline(y=0, line=dict(color=_GRID, width=1), row=2, col=1)

        # ── Row 3: cumulative drawdown ─────────────────────────
        equity_arr = np.array(equity)
        peak       = np.maximum.accumulate(equity_arr)
        drawdown   = list(equity_arr - peak)
        fig.add_trace(
            go.Scatter(
                x=trade_nums,
                y=drawdown,
                mode="lines",
                line=dict(color=_RED, width=1.5),
                fill="tozeroy",
                fillcolor="rgba(239,83,80,0.15)",
                hovertemplate="Trade #%{x}<br>Drawdown: %{y:.2f}R<extra></extra>",
                name="Drawdown",
                showlegend=False,
            ),
            row=3, col=1,
        )
        fig.add_hline(y=0, line=dict(color=_GRID, width=1, dash="dot"), row=3, col=1)

        # ── Row 4: per-env summary table ──────────────────────
        import pandas as pd
        rows = [_env_summary_row(ei, pd.DataFrame(self._env_trades.get(ei, [])))
                for ei in env_ids]
        if rows:
            tbl_df = pd.DataFrame(rows)
            fig.add_trace(
                go.Table(
                    header=dict(
                        values=[f"<b>{c}</b>" for c in tbl_df.columns],
                        fill_color=_PAPER,
                        font=dict(color=_TEXT, size=10),
                        align="center",
                        line_color=_GRID,
                        height=24,
                    ),
                    cells=dict(
                        values=[tbl_df[c].tolist() for c in tbl_df.columns],
                        fill_color=_BG,
                        font=dict(color=_TEXT, size=10),
                        align=["center"] * len(tbl_df.columns),
                        line_color=_GRID,
                        height=20,
                    ),
                ),
                row=4, col=1,
            )

        # ── Layout ────────────────────────────────────────────
        n_wins  = sum(agg_is_win)
        wr_pct  = 100 * n_wins / n_agg if n_agg else 0.0
        total_r = sum(agg_pnl)
        clr     = _GREEN if total_r >= 0 else _RED

        title = (
            f"Fold {fold_id:02d} — Training Journal  |  "
            f"{n_agg} trades  |  WR {wr_pct:.0f}%  |  "
            f"<span style='color:{clr}'>{total_r:+.2f}R</span>"
        )
        fig.update_layout(
            title=dict(text=title, font=dict(size=14, color=_TEXT), x=0.5),
            paper_bgcolor=_PAPER,
            plot_bgcolor=_BG,
            font=dict(color=_TEXT, family="monospace"),
            showlegend=False,
            dragmode="pan",
            margin=dict(l=60, r=30, t=60, b=60),
            height=1200,
        )
        # Range slider on bottom subplot for horizontal scrolling
        fig.update_xaxes(
            gridcolor=_GRID, zeroline=False,
            rangeslider=dict(visible=False),
        )
        fig.update_xaxes(
            rangeslider=dict(visible=True, thickness=0.04, bgcolor=_PAPER),
            row=3, col=1,
        )
        fig.update_yaxes(gridcolor=_GRID, zeroline=False)

        path = fold_dir / f"fold_{fold_id:02d}_training_journal.html"
        fig.write_html(
            str(path),
            include_plotlyjs="cdn",
            config={"scrollZoom": True, "displayModeBar": True,
                    "modeBarButtonsToAdd": ["pan2d", "zoom2d"]},
        )
        if self.verbose:
            print(f"[FoldJournal] Chart  saved → {path}")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _env_summary_row(env_idx: int, df) -> dict:
    """Compute one summary row for a single env's trade DataFrame."""
    import pandas as pd

    if df is None or (hasattr(df, "empty") and df.empty) or len(df) == 0:
        return {
            "Env": f"{env_idx:02d}", "Trades": 0, "WR%": "—",
            "PnL_R": "—", "PF": "—", "Sharpe": "—",
            "AvgW_R": "—", "AvgL_R": "—", "AvgDur_min": "—",
        }

    n      = len(df)
    wins   = df[df["is_win"]]   if "is_win"   in df.columns else pd.DataFrame()
    losses = df[~df["is_win"]]  if "is_win"   in df.columns else pd.DataFrame()
    wr     = len(wins) / n if n else 0.0
    pnl_r  = float(df["pnl_r"].sum()) if "pnl_r" in df.columns else 0.0
    gw     = float(wins["pnl_r"].sum())    if len(wins)   else 0.0
    gl     = abs(float(losses["pnl_r"].sum())) if len(losses) else 1e-6
    pf     = min(gw / max(gl, 1e-6), 99.99)
    avg_w  = float(wins["pnl_r"].mean())    if len(wins)   else 0.0
    avg_l  = float(losses["pnl_r"].mean())  if len(losses) else 0.0
    avg_d  = float(df["duration_min"].mean()) if "duration_min" in df.columns else 0.0

    tr_arr = df["pnl_r"].values if "pnl_r" in df.columns else np.array([])
    if len(tr_arr) >= 5 and tr_arr.std() > 0.01:
        sharpe = float(np.clip(tr_arr.mean() / tr_arr.std(), -9.99, 9.99))
    else:
        sharpe = 0.0

    return {
        "Env":       f"{env_idx:02d}",
        "Trades":    n,
        "WR%":       f"{wr*100:.1f}%",
        "PnL_R":     f"{pnl_r:+.2f}",
        "PF":        f"{pf:.2f}",
        "Sharpe":    f"{sharpe:.2f}",
        "AvgW_R":    f"{avg_w:+.3f}",
        "AvgL_R":    f"{avg_l:+.3f}",
        "AvgDur_min":f"{avg_d:.0f}",
    }


def _style_trades_sheet(ws, df) -> None:
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    green_fill = PatternFill("solid", fgColor="C8E6C9")
    red_fill   = PatternFill("solid", fgColor="FFCDD2")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    is_win_col = None
    for i, cell in enumerate(ws[1], 1):
        if cell.value == "is_win":
            is_win_col = i
            break

    for row in ws.iter_rows(min_row=2):
        if is_win_col:
            win_val = row[is_win_col - 1].value
            fill = green_fill if win_val else red_fill
            for cell in row:
                cell.fill = fill

    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 20)


def _style_summary_sheet(ws, df) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 22)
